"""Utilities to help extract data from Excel Files

"""

# IMPORTS
# ---------------------------------------------------------------------------------------------------------------------
import re
import string
import zipfile
import pandas as pd
from openpyxl import load_workbook
from logging import getLogger


# LOGGING
# ----------------------------------------------------------------------------------------------------------------------
logger = getLogger('j_rep')


# GLOBALS
# ----------------------------------------------------------------------------------------------------------------------
_pattern = re.compile('(?P<column>[A-Z]+)(?P<row>[0-9]+)')
_table_reference_pattern = re.compile('(?P<location>.+)\[(?P<file_name>.+)\](?P<sheet_name>.+)!(?P<data_range>.+)')
# _pattern = re.compile('(?P<cs>[A-Z]+)(?P<rs>[0-9]+):(?P<ce>[A-Z]+)(?P<re>[0-9]+)')


# FUNCTIONS
# ---------------------------------------------------------------------------------------------------------------------
def name_to_number(column_name):
    """Returns the corresponding number of the given alphabetical sequence.
        E.g.: 'Z' -> 26, 'AA' -> 27
        MAX = XFD

    :param str column_name:
    :rtype : int
    """
    number = 0
    for c in column_name:
        if c in string.ascii_letters:
            number = number * 26 + (ord(c.upper()) - ord('A')) + 1
        else:
            raise Exception('column_name must be an ascii letter, "{}" is not.'.format(column_name))
    return number


def number_to_name(column_number):
    """Converts Excel column number to name.
        E.g.: 1  -> 'A', 28 -> 'AB'

    :param int column_number:
    :rtype : str
    """
    assert column_number >= 0

    dividend = column_number
    column_name = ''

    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_name = chr(65 + modulo) + column_name
        dividend = (dividend - modulo) // 26

    return column_name


def cell_to_row_column(cell_name):
    """Convert cell range string in A1 notation to numeric row/col pair.
        Returns: (row, col)

    :param str cell_range:
    :rtype : (int, int)
    """
    rc = [m.groupdict() for m in _pattern.finditer(cell_name)][0]
    return int(rc['row']), name_to_number(rc['column'])


def range_to_row_column(range_name):
    """Convert range string in A1:B3 notation to numeric row/col pair.
        Returns: (row1, col1), (row2, col2)

    :param str cell_range:
    :rtype : (int, int), (int, int)
    """
    return (cell_to_row_column(rn) for rn in range_name.split(':'))


def row_column_to_cell(rc_tuple):
    """Convert RC pair to range name
        E.g: (5, 2) -> 'B5'
    :param (int, int) rc_tuple:
    :rtype : str
    """
    return number_to_name(rc_tuple[1]) + str(rc_tuple[0])


def row_column_to_range(rc_start, rc_end):
    """Convert RC pair to range name
        E.g: (1, 1), (5, 2) -> 'A1:B5'
    """
    return '{}:{}'.format(row_column_to_cell(rc_start), row_column_to_cell(rc_end))


def decode_table_reference(table_reference):
    """Decode table reference
        E.g.: /path/to/[workbook.xls]Sheet!A1:B2 ->
        return {
            'location': '/path/to/',
            'file_name': 'workbook.xls',
            'sheet_name': 'Sheet',
            'data_range': 'A1:B2',
        }

    :param str table_reference:
    :return dict:
    """
    return [m.groupdict() for m in _table_reference_pattern.finditer(table_reference)][0]


def get_range(xl_sheet, range_name, values=True, truncate=False):
    """Reads data from specified range into a list
        Returns a list of either cell values or xlrd cell objects,

    :param xlrd.Sheet xl_sheet: Excel sheet
    :param str range_name: The Excel range name, e.g.: 'A1:G7'
    :param bool values: If true, returns a list of cell values. If false, returns a list of xlrd cell objects
    :rtype : list
    """
    from_range, to_range = range_to_row_column(range_name)
    range_cells = []

    for idx in range(from_range[0] - 1, to_range[0]):
        range_cells.append(xl_sheet.row_slice(idx, start_colx=from_range[1] - 1, end_colx=to_range[1]))

    if values:
        range_values = []
        for row in range_cells:
            if row[0] or not truncate:
                range_values.append([cell.value for cell in row])
        return range_values

    return range_cells


def get_table_header(header, row=2):
    """Standardize column names from either 1 or 2 rows.

    :param list header:
    :param int row:
    :rtype : list
    """
    if row == 1:
        return list(re.sub('\W', '', v.title()) if v else 'Unnamed{}'.format(k + 1) for k, v in enumerate(header))

    if row == 2:
        column_names = []
        for i in range(len(header[0])):
            if header[0][i]:
                if header[1][i]:
                    cn = '{}_{}'.format(header[0][i], header[1][i])
                else:
                    cn = header[0][i]
            elif header[1][i] and header[1][i - 1]:
                cn = '{}_{}'.format(header[0][i - 1], header[1][i])
            else:
                cn = 'Unnamed{}'.format(i + 1)

            column_names.append(re.sub('\W', '', cn.title()))

        return column_names


def get_column_types(type_table, column_types_override=None):
    """Automatically detect column types

    """
    type_names = ['Empty', 'String', 'Decimal', 'Datetime', 'Text', 'Integer']
    column_types = {}
    for k, v in type_table.iteritems():
        v = v.mask(v == 0)
        v = v.dropna()
        if not v.empty:
            likely_type = v.value_counts().idxmax()
        else:
            likely_type = 0
        column_types.update({k: type_names[int(likely_type)]})
    if column_types_override:
        for k, v in column_types_override.items():
            column_types.update({k: v})
    return column_types


def get_tables_from_sheet(xl_sheet, context, keyword: str = '.+', 
    header_offset: int = 0, header_rows: int = 1, table_width: int = 0, max_rows: int = 300, probe_col: int = 1, grace_rows: int = 6):
    """ Given a sheet, returns all table found from that sheet.

    """
    tables = []

    probe_range = row_column_to_range((1, probe_col), (max_rows, probe_col))
    # print(probe_range)

    # probe = xl_sheet.iter_rows(range_string='A1:A{max_rows}'.format(max_rows=max_rows))
    probe = xl_sheet.iter_rows(range_string=probe_range)
    probe = [i[0].value for i in probe]

    regex = re.compile(keyword)
    idx = []
    for i, j in enumerate(probe):
        if len(re.findall(regex, str(j))) > 0:
            idx.append(i)

    for i in idx:
        ds = de = i + header_offset + header_rows - 1
        # This breaks

        while ''.join([x for x in probe[de: de + 1 + grace_rows] if x is not None]) and (de < (len(probe)-2)):
            # print(''.join([x for x in probe[de: 1 + grace_rows] if x is not None])[:4])
            de += 1

        if de - ds > 0:
            data_range = row_column_to_range((i + header_offset + 1, 1), (de + 1, table_width))
            table = dict(description=probe[i].strip(), header_rows=header_rows, data_range=data_range)

            if context:
                table.update(context)
            tables.append(table)
        else:
            logger.debug('Dropped empty table in context: %s' % context['file_path'])

    return tables


def get_tables(df_files, sheet_filter, table_start_keyword, 
    header_offset: int = 0, header_rows: int = 1, table_width: int = 0, probe_col: int = 1, grace_rows: int = 6):
    """ Get all tables

    """
    sheet_pattern = re.compile(sheet_filter)
    tables = []

    for file_id, file_data in df_files.iterrows():
        try:
            workbook = load_workbook(filename=file_data['file_path'], use_iterators=True, data_only=True)
        except zipfile.BadZipFile:
            logger.warning('Corrupt file: %s' % file_data['file_path'])
        except PermissionError:
            logger.warning('Permission denied: %s' % file_data['file_path'])
        else:
            logger.debug('Opened: {}..{}'.format(file_data['file_path'][:15], file_data['file_path'][-50:]))
            sheets = (sheet for sheet in workbook.get_sheet_names() if len(re.findall(sheet_pattern, sheet)) > 0)

            for sheet_name in sheets:
                xl_sheet = workbook.get_sheet_by_name(sheet_name)
                context = dict(file_id=file_id, file_path=file_data['file_path'], sheet_name=sheet_name)
                tables += get_tables_from_sheet(
                    xl_sheet, context, keyword=table_start_keyword, 
                    header_offset=header_offset, header_rows=header_rows, table_width=table_width, probe_col=probe_col, grace_rows=grace_rows)

    if tables:
        df = pd.DataFrame(tables)
        df.index.name = 'id_table'
        logger.info('Found {} tables'.format(df.shape[0]))
        return df


def type_guess(cell):

    if cell.value is None:
        guess = 0
    elif cell.data_type == 'n':
        if cell.is_date:
            guess = 3
        else:
            guess = 2
    elif cell.data_type == 's':
        guess = 1
    else:
        guess = 4

    return guess


def open_xl_workbook(file_path: str):
    try:
        workbook = load_workbook(filename=file_path, use_iterators=True, data_only=True)
    except zipfile.BadZipFile:
        logger.warning('Corrupt file: %s' % file_path)
    except PermissionError:
        logger.warning('Permission denied: %s' % file_path)
    else:
        logger.info('Opened: {}...{}'.format(file_path[:15], file_path[-50:]))
        return workbook

    return None


def get_excel_data(tables, sheet_config, mappings, meta_range):
    """

    :param pandas.DataFrame tables:
    :return dict:
    """
    table_data_list = []
    table_type_list = []

    files = list(tables.file_path.unique())

    for file_path in files:

        workbook = open_xl_workbook(file_path)
        if workbook is None:
            break

        sheets = list(tables[tables.file_path == file_path].sheet_name.unique())

        for sheet_name in sheets:
            table_ranges = tables[(tables.file_path == file_path) & (tables.sheet_name == sheet_name)]
            xl_sheet = workbook.get_sheet_by_name(sheet_name)

            for table_id, table in table_ranges.iterrows():
                data = xl_sheet.iter_rows(range_string=table.data_range)
                cells = pd.DataFrame.from_records(data)

                header = sheet_config['header_rows']
                table_values = cells.loc[header:, :].applymap(lambda x: x.value)
                table_typing = cells.loc[header:, :].applymap(type_guess)

                # table_header = cells.loc[:header - 1, :].applymap(lambda x: x.value)
                # column_names = get_table_header(header)

                column_names = mappings['columns']
                table_values.columns = column_names
                table_typing.columns = column_names
                df = table_values

                df['table_description'] = table['description']
                df['table_id'] = table_id
                df['file_id'] = table['file_id']

                get_meta_data = True
                if get_meta_data:
                    meta_data = xl_sheet.iter_rows(range_string=meta_range)
                    meta_cells = pd.DataFrame.from_records(meta_data)

                    df['date'] = meta_cells.values[0][0].value

                df.index.name = 'row_number'
                df.reset_index(inplace=True)

                table_data_list.append(df)
                table_type_list.append(pd.DataFrame(data=table_typing, columns=column_names))

    data = pd.concat(table_data_list)
    data.reset_index(inplace=True)

    type_info = pd.concat(table_type_list, ignore_index=True)
    type_info.fillna(0, inplace=True)

    return {'values': data, 'type_info': type_info}


# TESTS
# ---------------------------------------------------------------------------------------------------------------------
import unittest


class TestExcel(unittest.TestCase):

    def setUp(self):
        self.column_number = 14
        self.range_name = 'AA6:B4'

    def test_column_conversion(self):
        assert self.column_number == name_to_number(number_to_name(self.column_number))

    def test_range_conversion(self):
        f, t = range_to_row_column(self.range_name)
        assert self.range_name == row_column_to_range(f, t)

    def test_tables(self):
        assert True

    def test_header_parse(self):
        header_test = [['id', 'delay', ''], ['', 'name', 'time']]
        header = get_table_header(header_test, row=2)

        assert header[1] == 'id'


def run_tests():
    suite = unittest.TestLoader().loadTestsFromTestCase(TestExcel)
    unittest.TextTestRunner(verbosity=2).run(suite)
